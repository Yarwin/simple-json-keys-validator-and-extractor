import json
import os
import openpyxl
from openpyxl.utils import get_column_letter


def get_path_and_check_if_json_exist(address):
    file_abs_path = os.path.abspath(os.path.dirname(__file__))
    dir_path = os.path.join(file_abs_path, 'DATA')
    json_file_path = os.path.join(dir_path, '{}.json'.format(address))
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)
    if not os.path.isfile(json_file_path):
        with open(json_file_path, 'w') as f:
            json.dump({}, f)

    return json_file_path


def write_data_in_row(data, sheet, interface, row):
    for key, entry in data.items():
        if type(entry) is dict:
            write_data_in_row(entry, sheet, interface, row)
        if type(entry) is list:
            sheet.cell(column=interface[key], row=row, value="; ".join(entry), auto_size = True)
            sheet.column_dimensions[get_column_letter(interface[key])].width = 60
        elif key not in interface:
            continue
        else:
            sheet.cell(column=interface[key], row=row, value=entry)


def write_data(data, sheet, interface):
    for key in interface.keys():
        sheet.cell(column=interface[key], row=1, value=key)
    for i, key in enumerate(data.keys()):
        sheet.cell(column=1, row=i+2, value=key)
        write_data_in_row(data[key], sheet, interface, i+2)


def write_to_excel(data, interface, output):
    wb = openpyxl.Workbook()
    sheet = wb.active
    write_data(data, sheet, interface)
    wb.save(output)

